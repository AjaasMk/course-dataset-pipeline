from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

from .domains import normalize_discipline

CATEGORY_HEADER = "category"
COURSE_HEADER = "course name"
MAX_SLUG_LENGTH = 80
MAX_COURSE_ID_LENGTH = 64


class CourseListError(RuntimeError):
    pass


@dataclass(frozen=True)
class CourseEntry:
    course_id: str
    slug: str
    course_name: str
    category: str
    categories: tuple[str, ...]
    discipline: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "course_id": self.course_id,
            "slug": self.slug,
            "course_name": self.course_name,
            "category": self.category,
            "categories": list(self.categories),
            "discipline": self.discipline,
        }


def slugify(value: str) -> str:
    cleaned = "".join(char.lower() if char.isalnum() else "-" for char in value)
    return "-".join(part for part in cleaned.split("-") if part)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("—", "-").replace("–", "-").replace("’", "'")
    return re.sub(r"\s+", " ", text).strip()


def _read_sheet(path: Path) -> list[tuple[str, str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise CourseListError("openpyxl is required to read the course list workbook") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise CourseListError(f"{path} is empty")

    header = [_clean(cell).lower() for cell in rows[0]]
    try:
        category_index = header.index(CATEGORY_HEADER)
        course_index = header.index(COURSE_HEADER)
    except ValueError as exc:
        raise CourseListError(
            f"{path} must have '{CATEGORY_HEADER}' and '{COURSE_HEADER}' headers, found {header}"
        ) from exc

    pairs: list[tuple[str, str]] = []
    for row in rows[1:]:
        category = _clean(row[category_index]) if len(row) > category_index else ""
        course = _clean(row[course_index]) if len(row) > course_index else ""
        if category and course:
            pairs.append((category, course))
    return pairs


def load_course_list(path: Path | str) -> list[CourseEntry]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise CourseListError(f"course list not found: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise CourseListError(f"expected an .xlsx course list, got {workbook_path.suffix!r}")

    pairs = _read_sheet(workbook_path)
    if not pairs:
        raise CourseListError(f"{workbook_path} produced no course rows")

    order: list[str] = []
    categories_by_course: dict[str, list[str]] = {}
    for category, course in pairs:
        if course not in categories_by_course:
            categories_by_course[course] = []
            order.append(course)
        if category not in categories_by_course[course]:
            categories_by_course[course].append(category)

    entries: list[CourseEntry] = []
    used_slugs: set[str] = set()
    for course in order:
        categories = tuple(categories_by_course[course])
        slug = slugify(course)[:MAX_SLUG_LENGTH].rstrip("-")
        if not slug or slug in used_slugs:
            raise CourseListError(f"course {course!r} does not produce a unique slug")
        used_slugs.add(slug)

        course_id = f"crs_{slug.replace('-', '_')}"[:MAX_COURSE_ID_LENGTH].rstrip("_")
        entries.append(
            CourseEntry(
                course_id=course_id,
                slug=slug,
                course_name=course,
                category=categories[0],
                categories=categories,
                discipline=normalize_discipline(categories[0]),
            )
        )

    ids = [entry.course_id for entry in entries]
    if len(set(ids)) != len(ids):
        raise CourseListError("course ids are not unique after truncation")
    return entries


def categories(entries: list[CourseEntry]) -> list[str]:
    seen: list[str] = []
    for entry in entries:
        for category in entry.categories:
            if category not in seen:
                seen.append(category)
    return seen


def iter_entries(
    entries: list[CourseEntry],
    *,
    discipline: str | None = None,
    limit: int | None = None,
) -> Iterator[CourseEntry]:
    wanted = normalize_discipline(discipline) if discipline else ""
    count = 0
    for entry in entries:
        if wanted and wanted not in {normalize_discipline(c) for c in entry.categories}:
            continue
        yield entry
        count += 1
        if limit is not None and count >= limit:
            return
