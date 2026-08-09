import re
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl
from pydantic import BaseModel

DEFAULT_TAXONOMY_PATH = Path("docs/COURSE LIST.xlsx")

_BULLET = "•"
_SLUG_STRIP = re.compile(r"[^a-z0-9]+")


class Course(BaseModel):
    course_id: str
    standard_course_name: str
    fields: list[str]
    aliases: list[str]


def _slugify(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return _SLUG_STRIP.sub("_", normalized.lower()).strip("_")


def _parse_aliases(cell: Optional[str]) -> list[str]:
    if not cell:
        return []

    text = str(cell)
    parts = (
        [line.strip().lstrip(_BULLET) for line in text.split("\n")]
        if _BULLET in text
        else text.split(",")
    )
    return [part.strip() for part in parts if part.strip()]


def load_taxonomy(path: Optional[Path] = None) -> list[Course]:
    resolved = path if path is not None else DEFAULT_TAXONOMY_PATH
    workbook = openpyxl.load_workbook(resolved, data_only=True, read_only=True)

    merged: dict[str, Course] = {}
    for sheet_name in workbook.sheetnames:
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            name = str(row[0]).strip() if row and row[0] else ""
            if not name:
                continue

            aliases = _parse_aliases(row[1] if len(row) > 1 else None)
            key = name.lower()
            existing = merged.get(key)
            if existing is None:
                merged[key] = Course(
                    course_id=_slugify(name),
                    standard_course_name=name,
                    fields=[sheet_name],
                    aliases=aliases,
                )
                continue

            if sheet_name not in existing.fields:
                existing.fields.append(sheet_name)
            for alias in aliases:
                if alias not in existing.aliases:
                    existing.aliases.append(alias)

    workbook.close()
    return list(merged.values())
